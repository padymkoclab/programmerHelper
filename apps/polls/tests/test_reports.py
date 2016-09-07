
import collections
import itertools
import io
from unittest import mock
from unittest import skip

from django.core.management import call_command

import pytest
from PyPDF2 import PdfFileReader
import openpyxl

from utils.django.datetime_utils import (
    convert_date_to_django_date_format,
    get_current_timezone_offset,
    get_year_by_slavic_aryan_calendar
)
from utils.django.utils import get_location
from utils.django.test_utils import EnhancedTestCase

from config.admin import AdminSite
from apps.polls.models import Poll, Vote
from apps.polls.admin import PollAdmin
from apps.polls.factories import PollFactory, ChoiceFactory


class PollPDFReportWithoutObjectsTests(EnhancedTestCase):
    """
    Test for the PollPDFReport considering that polls, choices and votes
    does not exists yet.
    """

    @classmethod
    def setUpTestData(cls):

        # an active superuser for access to the admin
        call_command('factory_test_users', '1')
        cls.active_superuser = cls.django_user_model.objects.get()
        cls._make_user_as_active_superuser(cls.active_superuser)

        # full name of the user for check up in generated reports
        cls.superuser_full_name = cls.active_superuser.get_full_name()

        # django admin class
        cls.PollAdmin = PollAdmin(Poll, AdminSite)

        # url for make reports in PDF
        cls.url = cls.reverse('admin:polls_make_report')

    def _get_text_for_empty_page_polls(self, number_page_in_doc):
        """Return text for page of polls considering number page if polls does not exists yet."""

        text = 'Report about polls\nPolls\nPage {0}\nSubject: Polls\nPolls are not exists yet\n'
        return text.format(number_page_in_doc)

    def _get_text_for_empty_page_choices(self, number_page_in_doc):
        """Return text for page of choices considering number page if choices does not exists yet."""

        text = 'Report about polls\nChoices\nPage {0}\nSubject: Choices\nChoices are not exists yet\n'
        return text.format(number_page_in_doc)

    def _get_text_for_empty_page_votes(self, number_page_in_doc):
        """Return text for page of votes considering number page if votes does not exists yet."""

        text = ''.join([
            'Report about polls\nVotes\nPage {0}\nSubject: Votes\nVotes are not exists yet\n'.format(
                number_page_in_doc
            ),
            '\n'.join([
                'Oct 2015',
                'Nov 2015',
                'Dec 2015',
                'Jan 2016',
                'Feb 2016',
                'Mar 2016',
                'Apr 2016',
                'May 2016',
                'Jun 2016',
                'Jul 2016',
                'Aug 2016',
                'Sep 2016',
            ]),
            '\n0 votes \n1 votes \n2 votes \n3 votes \n4 votes \n5 votes \n',
            '6 votes \n7 votes \n8 votes \n9 votes \n10 votes \n',
            '0\n0\n0\n0\n0\n0\n0\n0\n0\n0\n0\n0\nCount votes for the past year\n',
        ])
        return text.format(number_page_in_doc)

    def _get_text_for_empty_page_voters(self, number_page_in_doc):
        """Return text for page of voters considering number page if votes does not exists yet."""

        text = 'Report about polls\nVoters\nPage {0}\nSubject: Voters\nVotes are not exists yet\n'
        return text.format(number_page_in_doc)

    def _get_text_for_empty_page_results(self, number_page_in_doc):
        """Return text for page of results of polls considering number page if polls does not exists yet."""

        text = 'Report about polls\nResults of polls\nPage {0}\nSubject: Results of polls\nPolls are not exists yet\n'
        return text.format(number_page_in_doc)

    def _test_doc_info(self, combination):
        """Tests for an each generated report-document."""

        # count pages must be +2, because always presents is page Title, Content, Statistics
        # and in self.data is type output as PDF
        count_pages = len(combination) + 3

        doc_info = self.doc.getDocumentInfo()
        self.assertEqual(self.doc.getNumPages(), count_pages)
        self.assertEqual(doc_info['/Creator'], self.settings.SITE_NAME)
        self.assertEqual(doc_info['/Keywords'], 'Polls, votes, voters')
        self.assertEqual(doc_info['/Subject'], self.subject)
        self.assertEqual(doc_info['/Title'], 'Report about polls')
        self.assertEqual(doc_info['/Author'], self.superuser_full_name)

    def _tests_title_page(self, now, request):
        """Tests for a first (title) page of an each generated report."""

        page0Text = self.doc.getPage(0).extractText()

        self.assertIn(self.settings.SITE_NAME, page0Text)
        self.assertIn(
            '{0} A. D. ({1}'.format(
                now.year,
                get_year_by_slavic_aryan_calendar(now)
            ),
            page0Text
        )
        self.assertIn('Report', page0Text)
        self.assertIn('"Report about polls"', page0Text)

        location = get_location(request)
        if location is None:
            location = '(Not possible determinate location)'
        self.assertIn('Location: (Not possible determinate location)'.format(location), page0Text)
        self.assertIn('Author: {0}'.format(self.superuser_full_name), page0Text)
        self.assertIn(
            'Generated: {0}'.format(convert_date_to_django_date_format(now)),
            page0Text
        )
        self.assertIn(
            'Timezone: Europe/Kiev ({1})'.format(
                self.settings.TIME_ZONE,
                get_current_timezone_offset()
            ),
            page0Text
        )

    def _tests_content_page(self):
        """Tests for a second page in an each generated report."""

        page1Text = self.doc.getPage(1).extractText()
        self.assertEqual(
            ''.join([
                'Report about polls\nContent\nPage 2\nContent of report\n',
                'This report contains a next subject:\n{0}\n'.format(self.lst_themes),
            ]),
            page1Text
        )

    def _test_statistics_page(self):
        """Tests for a third page - statistics of an each generated report."""

        page2Text = self.doc.getPage(2).extractText()
        self.assertEqual(
            ''.join([
                'Report about polls\nStatistics\nPage 3\nStatictics\nCount polls\n0\nCount opened polls\n0\n',
                'Count closed polls\n0\nCount draft polls\n0\nCount choices\n0\nCount votes\n0\nCount voters\n0\n',
                'Average a count votes in the polls\n0.0\nAverage a count choices in the polls\n0.0\n',
                'Date a latest vote\n\nA latest voter\n\nA poll with the latest vote\n\nA latest selected choice\n\n'
            ]),
            page2Text
        )

    @mock.patch('django.utils.timezone.now')
    def _get_doc_from_request_and_now_and_request(self, mock_now):
        """Return a generated PDF in in-memory file, now datetime and request."""

        # mock current datetime
        mock_now.return_value = self.timezone.datetime(2016, 9, 22, tzinfo=self.timezone.utc)

        request = self.factory.post(self.url, self.data)
        request.user = self.active_superuser

        now = self.timezone.now()

        response = self.PollAdmin.view_make_report(request)

        # in-memory file for PDF response
        doc = PdfFileReader(io.BytesIO((response.getvalue())))

        return doc, now, request

    def test_generated_reports_on_different_themes_if_no_polls_and_choices_and_votes_at_all(self):

        # all possible themes
        themes = {
            'polls': 'polls',
            'choices': 'choices',
            'votes': 'votes',
            'voters': 'voters',
            'results': 'results',
        }

        # create all possible combinations from the themes
        combinations = list()
        for i in range(1, len(themes) + 1):
            combinations.extend(map(dict, itertools.combinations(themes.items(), i)))

        # names of the themes and their functions for empty text on their pages
        themes_and_functions = {
            'polls': self._get_text_for_empty_page_polls,
            'choices': self._get_text_for_empty_page_choices,
            'votes': self._get_text_for_empty_page_votes,
            'voters': self._get_text_for_empty_page_voters,
            'results': self._get_text_for_empty_page_results,
        }

        # themes and their priorities in the report
        themes_and_priorities = collections.OrderedDict([
            ('polls', 1),
            ('choices', 2),
            ('votes', 3),
            ('voters', 4),
            ('results', 5),
        ])

        for combination in combinations:

            ordered_themes_and_function = collections.OrderedDict(
                (k, themes_and_functions[k]) for k, v in themes_and_priorities.items()
                if k in combination
            )

            # needed data of a POST request for generate PDF report
            self.data = {'output_report': 'report_pdf'}

            # add combination of themes for POST data
            self.data.update(combination)

            # determinated a subject of the report and it`s themes
            subject = list()
            lst_themes = list()
            if 'polls' in self.data:
                subject.append('Polls')
                lst_themes.append('Polls')
            if 'choices' in self.data:
                subject.append('Choices')
                lst_themes.append('Choices')
            if 'votes' in self.data:
                subject.append('Votes')
                lst_themes.append('Votes')
            if 'voters' in self.data:
                subject.append('Voters')
                lst_themes.append('Voters')
            if 'results' in self.data:
                subject.append('Results of polls')
                lst_themes.append('Results of polls')
            self.subject = ', '.join(subject).capitalize()
            self.lst_themes = '\n'.join(lst_themes)

            # get output of PDF report, now and request
            self.doc, now, request = self._get_doc_from_request_and_now_and_request()

            # tests basic of the report
            self._test_doc_info(combination)

            # tests for a title page on based the now datetime and the request
            self._tests_title_page(now, request)

            # tests for a page with content of the report
            self._tests_content_page()

            # tests for a page with statistics of the report
            self._test_statistics_page()

            # a number of page started from 2
            i = 2
            for func in ordered_themes_and_function.values():

                # number page in PDF document, numeration from 0
                page = i + 1

                # number page in report, numeration from 1
                number_page_in_doc = page + 1

                self.assertEqual(func(number_page_in_doc=number_page_in_doc), self.doc.getPage(page).extractText())
                i += 1


@skip('PyPDF does not support for unicode')
class PollPDFReportWithObjectsTests(EnhancedTestCase):
    """
    Test for the PollPDFReport considering that polls, choices and votes
    does not exists yet.
    """

    @classmethod
    def setUpTestData(cls):

        # an active superuser for access to the admin
        call_command('factory_test_users', '4')
        cls.active_superuser, cls.user1, cls.user2, cls.user3 = cls.django_user_model.objects.all()
        cls._make_user_as_active_superuser(cls.active_superuser)

        # full name of the user for check up in generated reports
        cls.superuser_full_name = cls.active_superuser.get_full_name()

        # generate polls, choices and votes
        cls.poll1 = PollFactory(status='opened')
        cls.choice11 = ChoiceFactory(poll=cls.poll1)
        cls.choice12 = ChoiceFactory(poll=cls.poll1)
        cls.choice13 = ChoiceFactory(poll=cls.poll1)

        cls.poll2 = PollFactory(status='draft')
        cls.choice21 = ChoiceFactory(poll=cls.poll2)
        cls.choice22 = ChoiceFactory(poll=cls.poll2)

        cls.poll3 = PollFactory(status='draft')
        cls.choice31 = ChoiceFactory(poll=cls.poll3)

        cls.poll4 = PollFactory(status='opened')
        cls.poll5 = PollFactory(status='opened')

        Vote.objects.create(user=cls.active_superuser, poll=cls.poll1, choice=cls.choice11)
        Vote.objects.create(user=cls.user1, poll=cls.poll1, choice=cls.choice11)
        Vote.objects.create(user=cls.user1, poll=cls.poll2, choice=cls.choice21)
        Vote.objects.create(user=cls.user2, poll=cls.poll1, choice=cls.choice21)
        Vote.objects.create(user=cls.user2, poll=cls.poll2, choice=cls.choice22)
        Vote.objects.create(user=cls.active_superuser, poll=cls.poll2, choice=cls.choice21)
        cls.vote = Vote.objects.create(user=cls.user2, poll=cls.poll3, choice=cls.choice31)

        # django admin class
        cls.PollAdmin = PollAdmin(Poll, AdminSite)

        # url for make reports in PDF
        cls.url = cls.reverse('admin:polls_make_report')

    def setUp(self):

        # needed data of a POST request for generate PDF report
        self.data = {'output_report': 'report_pdf'}

        # in-memory file for PDF response
        self.buffer = io.BytesIO()

    def _get_text_for_empty_page_polls(self, number_page_in_doc):
        """Return text for page of polls considering number page if polls does not exists yet."""

        text = 'Report about polls\nPolls\nPage {0}\nSubject: Polls\nPolls are not exists yet\n'
        return text.format(number_page_in_doc)

    def _get_text_for_empty_page_choices(self, number_page_in_doc):
        """Return text for page of choices considering number page if choices does not exists yet."""

        text = 'Report about polls\nChoices\nPage {0}\nSubject: Choices\nChoices are not exists yet\n'
        return text.format(number_page_in_doc)

    def _get_text_for_empty_page_votes(self, number_page_in_doc):
        """Return text for page of votes considering number page if votes does not exists yet."""

        text = ''.join([
            'Report about polls\nVotes\nPage {0}\nSubject: Votes\nVotes are not exists yet\n'.format(
                number_page_in_doc
            ),
            '\n'.join([
                'Oct 2015',
                'Nov 2015',
                'Dec 2015',
                'Jan 2016',
                'Feb 2016',
                'Mar 2016',
                'Apr 2016',
                'May 2016',
                'Jun 2016',
                'Jul 2016',
                'Aug 2016',
                'Sep 2016',
            ]),
            '\n0 votes \n1 votes \n2 votes \n3 votes \n4 votes \n5 votes \n',
            '6 votes \n7 votes \n8 votes \n9 votes \n10 votes \n',
            '0\n0\n0\n0\n0\n0\n0\n0\n0\n0\n0\n0\nCount votes for the past year\n',
        ])
        return text.format(number_page_in_doc)

    def _get_text_for_empty_page_voters(self, number_page_in_doc):
        """Return text for page of voters considering number page if votes does not exists yet."""

        text = 'Report about polls\nVoters\nPage {0}\nSubject: Voters\nVotes are not exists yet\n'
        return text.format(number_page_in_doc)

    def _get_text_for_empty_page_results(self, number_page_in_doc):
        """Return text for page of results of polls considering number page if polls does not exists yet."""

        text = 'Report about polls\nResults of polls\nPage {0}\nSubject: Results of polls\nPolls are not exists yet\n'
        return text.format(number_page_in_doc)

    def _test_doc_info(self, doc, subject, count_pages):
        """Tests for an each generated report-document."""

        doc_info = doc.getDocumentInfo()
        self.assertEqual(doc.getNumPages(), count_pages)
        self.assertEqual(doc_info['/Creator'], self.settings.SITE_NAME)
        self.assertEqual(doc_info['/Keywords'], 'Polls, votes, voters')
        self.assertEqual(doc_info['/Subject'], subject)
        self.assertEqual(doc_info['/Title'], 'Report about polls')
        self.assertEqual(doc_info['/Author'], self.superuser_full_name)

    def _tests_title_page(self, doc, now, request):
        """Tests for a first (title) page of an each generated report."""

        page0Text = doc.getPage(0).extractText()

        self.assertIn(self.settings.SITE_NAME, page0Text)
        self.assertIn(
            '{0} A. D. ({1}'.format(
                now.year,
                get_year_by_slavic_aryan_calendar(now)
            ),
            page0Text
        )
        self.assertIn('Report', page0Text)
        self.assertIn('"Report about polls"', page0Text)

        location = get_location(request)
        if location is None:
            location = '(Not possible determinate location)'
        self.assertIn('Location: (Not possible determinate location)'.format(location), page0Text)
        self.assertIn('Author: {0}'.format(self.superuser_full_name), page0Text)
        self.assertIn('Generated: {0}'.format(convert_date_to_django_date_format(now)), page0Text)
        self.assertIn(
            'Timezone: Europe/Kiev ({1})'.format(
                self.settings.TIME_ZONE,
                get_current_timezone_offset()
            ),
            page0Text
        )

    def _tests_content_page(self, doc, themes):
        """Tests for a second page in an each generated report."""

        page1Text = doc.getPage(1).extractText()
        listing_themes = '\n'.join(themes)
        self.assertEqual(
            ''.join([
                'Report about polls\nContent\nPage 2\nContent of report\n',
                'This report contains a next subject:\n{0}\n'.format(listing_themes),
            ]),
            page1Text
        )

    def _test_statistics_page(self, doc):

        page2Text = doc.getPage(2).extractText()
        self.assertEqual(
            ''.join([
                'Report about polls\nStatistics\nPage 3\nStatictics\nCount polls\n5\nCount opened polls\n3\n',
                'Count closed polls\n0\nCount draft polls\n2\nCount choices\n6\nCount votes\n7\nCount voters\n3\n',
                'Average a count votes in the polls\n1.4\nAverage a count choices in the polls\n1.2\n',
                'Date a latest vote\n{0}\nA latest voter\n{1}\n'.format(
                    convert_date_to_django_date_format(self.vote.date_voting),
                    self.user2.get_full_name(),
                ),
                'A poll with the latest vote\n{0}\nA latest selected choice\n{0}\n'.format(
                    self.vote.poll,
                    self.vote.choice,
                )
            ]),
            page2Text
        )

    @mock.patch('django.utils.timezone.now')
    def _get_doc_from_request_and_now_and_request(self, mock_now):
        """Return a generated PDF in in-memory file, now datetime and request."""

        # mock current datetime
        mock_now.return_value = self.timezone.datetime(2016, 9, 22, tzinfo=self.timezone.utc)

        request = self.factory.post(self.url, self.data)
        request.user = self.active_superuser

        now = self.timezone.now()

        response = self.PollAdmin.view_make_report(request)

        self.buffer.write(response.getvalue())

        doc = PdfFileReader(self.buffer)

        return doc, now, request

    @pytest.mark.skip('Does not working with unicode')
    def test_without_objects_on_theme_polls(self):

        self.data['polls'] = 'polls'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Polls', 5)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Polls'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_polls(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

    def test_without_objects_on_theme_choices(self):

        self.data['choices'] = 'choices'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Choices', 4)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Choices'])
        self._test_statistics_page(doc)

        # page of choices
        self.assertEqual(
            self._get_text_for_empty_page_choices(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

    def test_without_objects_on_theme_votes(self):

        self.data['votes'] = 'votes'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Votes', 4)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Votes'])
        self._test_statistics_page(doc)

        # page of votes
        self.assertEqual(
            self._get_text_for_empty_page_votes(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

    def test_without_objects_on_theme_voters(self):

        self.data['voters'] = 'voters'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Voters', 4)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Voters'])
        self._test_statistics_page(doc)

        # page of voters
        self.assertEqual(
            self._get_text_for_empty_page_voters(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

    def test_without_objects_on_theme_results(self):

        self.data['results'] = 'results'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Results of polls', 4)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Results of polls'])
        self._test_statistics_page(doc)

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_results(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

    def test_without_objects_on_theme_polls_and_choices(self):

        self.data['polls'] = 'polls'
        self.data['choices'] = 'choices'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Polls, choices', 5)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Polls', 'Choices'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_polls(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

        # page of choices
        self.assertEqual(
            self._get_text_for_empty_page_choices(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )

    def test_without_objects_on_theme_polls_and_votes(self):

        self.data['polls'] = 'polls'
        self.data['votes'] = 'votes'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Polls, votes', 5)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Polls', 'Votes'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_polls(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

        # page of votes
        self.assertEqual(
            self._get_text_for_empty_page_votes(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )

    def test_without_objects_on_theme_polls_and_voters(self):

        self.data['polls'] = 'polls'
        self.data['voters'] = 'voters'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Polls, voters', 5)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Polls', 'Voters'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_polls(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

        # page of voters
        self.assertEqual(
            self._get_text_for_empty_page_voters(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )

    def test_without_objects_on_theme_polls_and_results(self):

        self.data['polls'] = 'polls'
        self.data['results'] = 'results'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Polls, results of polls', 5)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Polls', 'Results of polls'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_polls(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_results(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )

    def test_without_objects_on_theme_choices_and_votes(self):

        self.data['choices'] = 'choices'
        self.data['votes'] = 'votes'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Choices, votes', 5)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Choices', 'Votes'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_choices(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_votes(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )

    def test_without_objects_on_theme_choices_and_voters(self):

        self.data['choices'] = 'choices'
        self.data['voters'] = 'voters'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Choices, voters', 5)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Choices', 'Voters'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_choices(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_voters(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )

    def test_without_objects_on_theme_choices_and_results(self):

        self.data['choices'] = 'choices'
        self.data['results'] = 'results'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Choices, results of polls', 5)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Choices', 'Results of polls'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_choices(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_results(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )

    def test_without_objects_on_theme_votes_and_voters(self):

        self.data['votes'] = 'votes'
        self.data['voters'] = 'voters'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Votes, voters', 5)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Votes', 'Voters'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_votes(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_voters(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )

    def test_without_objects_on_theme_votes_and_results(self):

        self.data['votes'] = 'votes'
        self.data['results'] = 'results'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Votes, results of polls', 5)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Votes', 'Results of polls'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_votes(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_results(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )

    def test_without_objects_on_theme_voters_and_results(self):

        self.data['voters'] = 'voters'
        self.data['results'] = 'results'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Voters, results of polls', 5)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Voters', 'Results of polls'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_voters(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_results(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )

    def test_without_objects_on_theme_polls_and_choices_and_results(self):

        self.data['polls'] = 'polls'
        self.data['results'] = 'results'
        self.data['choices'] = 'choices'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Polls, choices, results of polls', 6)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Polls', 'Choices', 'Results of polls'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_polls(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_choices(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_results(number_page_in_doc=6),
            doc.getPage(5).extractText()
        )

    def test_without_objects_on_theme_polls_and_choices_and_votes(self):

        self.data['polls'] = 'polls'
        self.data['votes'] = 'votes'
        self.data['choices'] = 'choices'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Polls, choices, votes', 6)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Polls', 'Choices', 'Votes'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_polls(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_choices(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_votes(number_page_in_doc=6),
            doc.getPage(5).extractText()
        )

    def test_without_objects_on_theme_polls_and_choices_and_voters(self):

        self.data['polls'] = 'polls'
        self.data['voters'] = 'voters'
        self.data['choices'] = 'choices'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Polls, choices, voters', 6)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Polls', 'Choices', 'Voters'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_polls(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_choices(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_voters(number_page_in_doc=6),
            doc.getPage(5).extractText()
        )

    def test_without_objects_on_theme_polls_and_votes_and_voters(self):

        self.data['polls'] = 'polls'
        self.data['voters'] = 'voters'
        self.data['votes'] = 'votes'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Polls, votes, voters', 6)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Polls', 'Votes', 'Voters'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_polls(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_votes(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_voters(number_page_in_doc=6),
            doc.getPage(5).extractText()
        )

    def test_without_objects_on_theme_polls_and_votes_results(self):

        self.data['polls'] = 'polls'
        self.data['votes'] = 'votes'
        self.data['results'] = 'results'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Polls, votes, results of polls', 6)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Polls', 'Votes', 'Results of polls'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_polls(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_votes(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_results(number_page_in_doc=6),
            doc.getPage(5).extractText()
        )

    def test_without_objects_on_theme_polls_and_voters_results(self):

        self.data['polls'] = 'polls'
        self.data['voters'] = 'voters'
        self.data['results'] = 'results'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Polls, voters, results of polls', 6)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Polls', 'Voters', 'Results of polls'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_polls(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_voters(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_results(number_page_in_doc=6),
            doc.getPage(5).extractText()
        )

    def test_without_objects_on_theme_choices_and_votes_and_voters(self):

        self.data['choices'] = 'choices'
        self.data['voters'] = 'voters'
        self.data['votes'] = 'votes'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Choices, votes, voters', 6)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Choices', 'Votes', 'Voters'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_choices(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_votes(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_voters(number_page_in_doc=6),
            doc.getPage(5).extractText()
        )

    def test_without_objects_on_theme_choices_and_votes_and_results(self):

        self.data['choices'] = 'choices'
        self.data['results'] = 'results'
        self.data['votes'] = 'votes'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Choices, votes, results of polls', 6)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Choices', 'Votes', 'Results of polls'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_choices(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_votes(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_results(number_page_in_doc=6),
            doc.getPage(5).extractText()
        )

    def test_without_objects_on_theme_choices_and_voters_and_results(self):

        self.data['choices'] = 'choices'
        self.data['voters'] = 'voters'
        self.data['results'] = 'results'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Choices, voters, results of polls', 6)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Choices', 'Voters', 'Results of polls'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_choices(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_voters(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_results(number_page_in_doc=6),
            doc.getPage(5).extractText()
        )

    def test_without_objects_on_theme_votes_and_voters_and_results(self):

        self.data['results'] = 'results'
        self.data['voters'] = 'voters'
        self.data['votes'] = 'votes'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Votes, voters, results of polls', 6)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Votes', 'Voters', 'Results of polls'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_votes(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_voters(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )

        # page of results
        self.assertEqual(
            self._get_text_for_empty_page_results(number_page_in_doc=6),
            doc.getPage(5).extractText()
        )

    def test_without_objects_on_theme_polls_and_choices_and_votes_and_voters(self):

        self.data['polls'] = 'polls'
        self.data['choices'] = 'choices'
        self.data['voters'] = 'voters'
        self.data['votes'] = 'votes'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Polls, choices, votes, voters', 7)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Polls', 'Choices', 'Votes', 'Voters'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_polls(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )
        self.assertEqual(
            self._get_text_for_empty_page_choices(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )
        self.assertEqual(
            self._get_text_for_empty_page_votes(number_page_in_doc=6),
            doc.getPage(5).extractText()
        )
        self.assertEqual(
            self._get_text_for_empty_page_voters(number_page_in_doc=7),
            doc.getPage(6).extractText()
        )

    def test_without_objects_on_theme_polls_and_choices_and_votes_and_results(self):

        self.data['polls'] = 'polls'
        self.data['choices'] = 'choices'
        self.data['results'] = 'results'
        self.data['votes'] = 'votes'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Polls, choices, votes, results of polls', 7)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Polls', 'Choices', 'Votes', 'Results of polls'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_polls(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )
        self.assertEqual(
            self._get_text_for_empty_page_choices(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )
        self.assertEqual(
            self._get_text_for_empty_page_votes(number_page_in_doc=6),
            doc.getPage(5).extractText()
        )
        self.assertEqual(
            self._get_text_for_empty_page_results(number_page_in_doc=7),
            doc.getPage(6).extractText()
        )

    def test_without_objects_on_theme_polls_and_choices_votes_and_voters(self):

        self.data['polls'] = 'polls'
        self.data['choices'] = 'choices'
        self.data['voters'] = 'voters'
        self.data['votes'] = 'votes'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Polls, choices, votes, voters', 7)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Polls', 'Choices', 'Votes', 'Voters'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_polls(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )
        self.assertEqual(
            self._get_text_for_empty_page_choices(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )
        self.assertEqual(
            self._get_text_for_empty_page_votes(number_page_in_doc=6),
            doc.getPage(5).extractText()
        )
        self.assertEqual(
            self._get_text_for_empty_page_voters(number_page_in_doc=7),
            doc.getPage(6).extractText()
        )

    def test_without_objects_on_theme_polls_and_choices_and_voters_and_results(self):

        self.data['polls'] = 'polls'
        self.data['choices'] = 'choices'
        self.data['results'] = 'results'
        self.data['voters'] = 'voters'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Polls, choices, voters, results of polls', 7)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Polls', 'Choices', 'Voters', 'Results of polls'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_polls(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )
        self.assertEqual(
            self._get_text_for_empty_page_choices(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )
        self.assertEqual(
            self._get_text_for_empty_page_voters(number_page_in_doc=6),
            doc.getPage(5).extractText()
        )
        self.assertEqual(
            self._get_text_for_empty_page_results(number_page_in_doc=7),
            doc.getPage(6).extractText()
        )

    def test_without_objects_on_theme_choices_and_votes_voters_and_results(self):

        self.data['votes'] = 'votes'
        self.data['choices'] = 'choices'
        self.data['results'] = 'results'
        self.data['voters'] = 'voters'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Choices, votes, voters, results of polls', 7)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Choices', 'Votes', 'Voters', 'Results of polls'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_choices(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )
        self.assertEqual(
            self._get_text_for_empty_page_votes(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )
        self.assertEqual(
            self._get_text_for_empty_page_voters(number_page_in_doc=6),
            doc.getPage(5).extractText()
        )
        self.assertEqual(
            self._get_text_for_empty_page_results(number_page_in_doc=7),
            doc.getPage(6).extractText()
        )

    def test_without_objects_on_theme_polls_and_choices_and_votes_and_voters_and_results(self):

        self.data['polls'] = 'polls'
        self.data['votes'] = 'votes'
        self.data['choices'] = 'choices'
        self.data['results'] = 'results'
        self.data['voters'] = 'voters'

        doc, now, request = self._get_doc_from_request_and_now_and_request()

        self._test_doc_info(doc, 'Polls, choices, votes, voters, results of polls', 8)
        self._tests_title_page(doc, now, request)
        self._tests_content_page(doc, ['Polls', 'Choices', 'Votes', 'Voters', 'Results of polls'])
        self._test_statistics_page(doc)

        self.assertEqual(
            self._get_text_for_empty_page_polls(number_page_in_doc=4),
            doc.getPage(3).extractText()
        )
        self.assertEqual(
            self._get_text_for_empty_page_choices(number_page_in_doc=5),
            doc.getPage(4).extractText()
        )
        self.assertEqual(
            self._get_text_for_empty_page_votes(number_page_in_doc=6),
            doc.getPage(5).extractText()
        )
        self.assertEqual(
            self._get_text_for_empty_page_voters(number_page_in_doc=7),
            doc.getPage(6).extractText()
        )
        self.assertEqual(
            self._get_text_for_empty_page_results(number_page_in_doc=8),
            doc.getPage(7).extractText()
        )


class PollExcelReportWithoutObjectsTests(EnhancedTestCase):

    @classmethod
    def setUpTestData(cls):
        # create active superuser
        cls.call_command('factory_test_users', '1')
        cls.active_superuser = cls.django_user_model.objects.first()
        cls.active_superuser_full_name = cls.active_superuser.get_full_name()

        cls._make_user_as_active_superuser(cls.active_superuser)

        # admin class for Polls
        cls.PollAdmin = PollAdmin(Poll, AdminSite)

    def _tests_basic_document(self):

        self.assertEqual(self.excelfile.get_active_sheet().title, 'Statistics')

        sheet_names = ['Statistics']
        subject = []
        if 'polls' in self.data:
            sheet_names.append('Polls')
            subject.append('Polls')
        if 'choices' in self.data:
            sheet_names.append('Choices')
            subject.append('Choices')
        if 'votes' in self.data:
            sheet_names.append('Votes')
            subject.append('Votes')
        if 'voters' in self.data:
            sheet_names.append('Voters')
            subject.append('Voters')
        if 'results' in self.data:
            sheet_names.append('Results')
            subject.append('Results of polls')
        subject = ', '.join(subject).capitalize()

        self.assertEqual(self.excelfile.get_sheet_names(), sheet_names)
        self.assertEqual(self.excelfile.properties.creator, self.active_superuser_full_name)
        self.assertEqual(self.excelfile.properties.description, 'Report created with help library XlsxWriter 0.8.7.')
        self.assertEqual(self.excelfile.properties.title, 'Report about polls')
        self.assertEqual(self.excelfile.properties.keywords, 'Polls, votes, voters')
        self.assertEqual(self.excelfile.properties.subject, subject)

    def _tests_for_sheet_statistics_if_no_polls_choices_and_votes(self):

        sheet_statistics = self.excelfile.get_sheet_by_name('Statistics')
        self.assertEqual(sheet_statistics.cell('A2').value, 'Statistics')
        self.assertTrue(sheet_statistics.cell('A2').font.bold)
        self.assertEqual(sheet_statistics.cell('A5').value, 'Common statistics')
        self.assertTrue(sheet_statistics.cell('A5').font.bold)
        self.assertEqual(sheet_statistics.cell('A6').value, 'Count polls')
        self.assertEqual(sheet_statistics.cell('B6').value, 0)
        self.assertEqual(sheet_statistics.cell('A7').value, 'Count choices')
        self.assertEqual(sheet_statistics.cell('B7').value, 0)
        self.assertEqual(sheet_statistics.cell('A8').value, 'Count votes')
        self.assertEqual(sheet_statistics.cell('B8').value, 0)
        self.assertEqual(sheet_statistics.cell('A9').value, 'Count voters')
        self.assertEqual(sheet_statistics.cell('B9').value, 0)
        self.assertEqual(sheet_statistics.cell('A10').value, 'Count opened\npolls')
        self.assertEqual(sheet_statistics.cell('B10').value, 0)
        self.assertEqual(sheet_statistics.cell('A11').value, 'Count closed\npolls')
        self.assertEqual(sheet_statistics.cell('B11').value, 0)
        self.assertEqual(sheet_statistics.cell('A12').value, 'Count draft\npolls')
        self.assertEqual(sheet_statistics.cell('B12').value, 0)
        self.assertEqual(sheet_statistics.cell('A13').value, 'Average count\nchoices in polls')
        self.assertEqual(sheet_statistics.cell('B13').value, 0)
        self.assertEqual(sheet_statistics.cell('A14').value, 'Average count\nvotes in polls')
        self.assertEqual(sheet_statistics.cell('B14').value, 0)
        self.assertEqual(sheet_statistics.cell('A16').value, 'Latest vote')
        self.assertTrue(sheet_statistics.cell('A16').font.bold)
        self.assertEqual(sheet_statistics.cell('A17').value, 'Votes are not exists yet')
        self.assertTrue(sheet_statistics.cell('A17').font.bold)

    def _tests_for_sheet_polls_if_no_polls(self):
        sheet_polls = self.excelfile.get_sheet_by_name('Polls')
        self.assertEqual(sheet_polls.cell('A2').value, 'Polls')
        self.assertTrue(sheet_polls.cell('A2').font.bold)
        self.assertEqual(sheet_polls.cell('A5').value, '№')
        self.assertEqual(sheet_polls.cell('B5').value, 'Id')
        self.assertEqual(sheet_polls.cell('C5').value, 'Title')
        self.assertEqual(sheet_polls.cell('D5').value, 'Slug')
        self.assertEqual(sheet_polls.cell('E5').value, 'Description')
        self.assertEqual(sheet_polls.cell('F5').value, 'Count\nvotes')
        self.assertEqual(sheet_polls.cell('G5').value, 'Count\nchoices')
        self.assertEqual(sheet_polls.cell('H5').value, 'Status')
        self.assertEqual(sheet_polls.cell('I5').value, 'Latest changed\nof status')
        self.assertEqual(sheet_polls.cell('J5').value, 'Date modified')
        self.assertEqual(sheet_polls.cell('K5').value, 'Date added')
        self.assertEqual(sheet_polls.cell('A6').value, 'Polls are not exists yet')
        self.assertTrue(sheet_polls.cell('A6').font.bold)

    def _tests_for_sheet_choices_if_no_choices(self):
        sheet_choices = self.excelfile.get_sheet_by_name('Choices')
        self.assertEqual(sheet_choices.cell('A2').value, 'Choices')
        self.assertTrue(sheet_choices.cell('A2').font.bold)
        self.assertEqual(sheet_choices.cell('A5').value, '№')
        self.assertEqual(sheet_choices.cell('B5').value, 'Id')
        self.assertEqual(sheet_choices.cell('C5').value, 'Text of choice')
        self.assertEqual(sheet_choices.cell('D5').value, 'Poll')
        self.assertEqual(sheet_choices.cell('E5').value, 'Count\nvotes')
        self.assertEqual(sheet_choices.cell('A6').value, 'Choices are not exists yet')
        self.assertTrue(sheet_choices.cell('A6').font.bold)

    def _tests_for_sheet_votes_if_no_votes(self):
        sheet_votes = self.excelfile.get_sheet_by_name('Votes')
        self.assertEqual(sheet_votes.cell('A2').value, 'Votes')
        self.assertTrue(sheet_votes.cell('A2').font.bold)
        self.assertEqual(sheet_votes.cell('A5').value, '№')
        self.assertEqual(sheet_votes.cell('B5').value, 'Id')
        self.assertEqual(sheet_votes.cell('C5').value, 'Voter')
        self.assertEqual(sheet_votes.cell('D5').value, 'Poll')
        self.assertEqual(sheet_votes.cell('E5').value, 'Choice')
        self.assertEqual(sheet_votes.cell('F5').value, 'Date\nvoting')
        self.assertEqual(sheet_votes.cell('A6').value, 'Votes are not exists yet')
        self.assertTrue(sheet_votes.cell('A6').font.bold)
        self.assertEqual(sheet_votes.cell('H2').value, 'Count votes for the past year')
        self.assertTrue(sheet_votes.cell('H2').font.bold)
        self.assertEqual(sheet_votes.cell('H3').value, 'Month, year')
        self.assertTrue(sheet_votes.cell('H3').font.bold)
        self.assertEqual(sheet_votes.cell('H4').value, 'Count votes')
        self.assertTrue(sheet_votes.cell('H4').font.bold)
        self.assertEqual(sheet_votes.cell('I3').value, 'Nov 2011')
        self.assertEqual(sheet_votes.cell('I4').value, 0)
        self.assertEqual(sheet_votes.cell('J3').value, 'Dec 2011')
        self.assertEqual(sheet_votes.cell('J4').value, 0)
        self.assertEqual(sheet_votes.cell('K3').value, 'Jan 2012')
        self.assertEqual(sheet_votes.cell('K4').value, 0)
        self.assertEqual(sheet_votes.cell('L3').value, 'Feb 2012')
        self.assertEqual(sheet_votes.cell('L4').value, 0)
        self.assertEqual(sheet_votes.cell('M3').value, 'Mar 2012')
        self.assertEqual(sheet_votes.cell('M4').value, 0)
        self.assertEqual(sheet_votes.cell('N3').value, 'Apr 2012')
        self.assertEqual(sheet_votes.cell('N4').value, 0)
        self.assertEqual(sheet_votes.cell('O3').value, 'May 2012')
        self.assertEqual(sheet_votes.cell('O4').value, 0)
        self.assertEqual(sheet_votes.cell('P3').value, 'Jun 2012')
        self.assertEqual(sheet_votes.cell('P4').value, 0)
        self.assertEqual(sheet_votes.cell('Q3').value, 'Jul 2012')
        self.assertEqual(sheet_votes.cell('Q4').value, 0)
        self.assertEqual(sheet_votes.cell('R3').value, 'Aug 2012')
        self.assertEqual(sheet_votes.cell('R4').value, 0)
        self.assertEqual(sheet_votes.cell('S3').value, 'Sep 2012')
        self.assertEqual(sheet_votes.cell('S4').value, 0)
        self.assertEqual(sheet_votes.cell('T3').value, 'Oct 2012')
        self.assertEqual(sheet_votes.cell('T4').value, 0)

    def _tests_for_sheet_voters_if_no_voters(self):
        sheet_voters = self.excelfile.get_sheet_by_name('Voters')
        self.assertEqual(sheet_voters.cell('A2').value, 'Voters')
        self.assertTrue(sheet_voters.cell('A2').font.bold)
        self.assertEqual(sheet_voters.cell('A5').value, '№')
        self.assertEqual(sheet_voters.cell('B5').value, 'Id')
        self.assertEqual(sheet_voters.cell('C5').value, 'Full name')
        self.assertEqual(sheet_voters.cell('D5').value, 'Count votes')
        self.assertEqual(sheet_voters.cell('E5').value, 'Latest vote')
        self.assertEqual(sheet_voters.cell('F5').value, 'Is active\nvoter?')
        self.assertEqual(sheet_voters.cell('G5').value, 'All votes')
        self.assertEqual(sheet_voters.cell('A6').value, 'Votes are not exists yet')
        self.assertTrue(sheet_voters.cell('A6').font.bold)

    def _tests_for_sheet_results_if_no_results(self):
        sheet_results = self.excelfile.get_sheet_by_name('Results')
        self.assertEqual(sheet_results.cell('A2').value, 'Results')
        self.assertTrue(sheet_results.cell('A2').font.bold)
        self.assertEqual(sheet_results.cell('A4').value, 'Polls are not exists yet')
        self.assertTrue(sheet_results.cell('A4').value)

    @mock.patch('django.utils.timezone.now')
    def test_generated_reports_on_different_themes_if_no_polls_and_choices_and_votes_at_all(self, mock_now):

        mock_now.return_value = self.timezone.datetime(2012, 10, 15, 14, 7, 11, tzinfo=self.timezone.utc)

        # possible themes
        themes = {
            'polls': 'polls',
            'choices': 'choices',
            'votes': 'votes',
            'voters': 'voters',
            'results': 'results',
        }

        # combinations possible themes
        combinations = list()
        for i in range(1, len(themes) + 1):
            combinations.extend(map(dict, itertools.combinations(themes.items(), i)))

        for combination in combinations:

            # a report must be in Excel
            self.data = {'output_report': 'report_excel'}

            # add combinations of the themes
            self.data.update(themes)

            # make a request for get the Excel report as output
            request = self.factory.post('admin:polls_make_report', self.data)
            request.user = self.active_superuser
            response = self.PollAdmin.view_make_report(request)
            self.excelfile = openpyxl.load_workbook(io.BytesIO(response.getvalue()))

            # tests for basic of generated document
            self._tests_basic_document()

            # tests sheet of Statistcs
            self._tests_for_sheet_statistics_if_no_polls_choices_and_votes()

            # tests for sheet of 'Statistics'
            if 'polls' in self.data:
                self._tests_for_sheet_polls_if_no_polls()

            # tests for sheet of 'Choices'
            if 'choices' in self.data:
                self._tests_for_sheet_choices_if_no_choices()

            # tests for sheet of 'Votes'
            if 'votes' in self.data:
                self._tests_for_sheet_votes_if_no_votes()

            # tests for sheet of 'Voters'
            if 'voters' in self.data:
                self._tests_for_sheet_voters_if_no_voters()

            # tests for sheet of 'Results'
            if 'results' in self.data:
                self._tests_for_sheet_results_if_no_results()
