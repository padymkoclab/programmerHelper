
import csv

from django.core.urlresolvers import reverse
from django.contrib.contenttypes.models import ContentType
from django.contrib import admin
from django.http import HttpResponseRedirect, HttpResponse
from django.utils.translation import ugettext_lazy as _
from django.core import serializers
from django.utils.encoding import smart_str

import openpyxl
from openpyxl.cell import get_column_letter


def export_as_json(modeladmin, request, queryset):
    """ """

    response = HttpResponse(content_type='application/json')
    serializers.serialize('json', queryset, stream=response)
    return response
export_as_json.short_description = _('Export model in JSON file')


def export_as_xml(modeladmin, request, queryset):
    """ """

    response = HttpResponse(content_type='application/xml')
    serializers.serialize('xml', queryset, stream=response)
    return response
export_as_xml.short_description = _('Export model in XML file')


def export_as_yaml(modeladmin, request, queryset):
    """ """

    response = HttpResponse(content_type='application/yaml')
    serializers.serialize('yaml', queryset, stream=response)
    return response
export_as_yaml.short_description = _('Export model in YAML file')


def export_as_xlsx(self, request, queryset):
    """Export seleted rows in admin in XLSX-format."""

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "MyModel"

    row_num = 0

    columns = [
        # (u"ID", 15),
        (u"Title", 70),
        (u"Description", 70),
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            # obj.pk,
            obj.title,
            obj.description,
        ]
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.style.alignment.wrap_text = True

    wb.save(response)
    return response
export_as_xlsx.short_description = _('Export in excel format')


def export_as_csv(self, request, queryset):
    """Export seleted rows in admin in CSV-format."""

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="somefilename.csv"'
    writer = csv.writer(response, csv.excel)
    response.write('\ufeff'.encode('utf8'))
    writer.writerow([
        smart_str("ID"),
        smart_str("Title"),
        smart_str("Description"),
    ])
    for obj in queryset:
        writer.writerow([
            smart_str(obj.pk),
            smart_str(obj.title),
            smart_str(obj.description),
        ])
    return response
export_as_csv.short_description = _('Export in CSV format')
