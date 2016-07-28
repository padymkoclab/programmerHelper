
import csv

from django.core import serializers
from django.utils.encoding import smart_str
from django.http import HttpResponse
from django.utils.translation import ugettext_lazy as _
from django.utils.translation import ungettext

import openpyxl
from openpyxl.cell import get_column_letter


def make_users_as_non_superuser(self, request, queryset):
    """Admin action by selected users will be non-superusers."""
    for obj in queryset:
        obj.is_superuser = False
        obj.save()
    message = ungettext(
        'Succefull update parametrs one user',
        'Succefull update parametrs %(count_users)d users',
        len(queryset)
    ) % {
        'count_users': len(queryset),
    }
    self.message_user(request, message)
make_users_as_non_superuser.short_description = _('Make selected users as non-superuser')


def make_users_as_superuser(self, request, queryset):
    """Admin action by selected users will be superusers."""
    for obj in queryset:
        obj.is_superuser = True
        obj.save()
    message = ungettext(
        'Succefull update parametrs one user',
        'Succefull update parametrs %(count_users)d users',
        len(queryset)
    ) % {
        'count_users': len(queryset),
    }
    self.message_user(request, message)
make_users_as_superuser.short_description = _('Make selected users as superuser')


def make_users_as_non_active(self, request, queryset):
    """Admin action by selected users will be non-superusers."""
    for obj in queryset:
        obj.is_active = False
        obj.save()
    message = ungettext(
        'Succefull update parametrs one user',
        'Succefull update parametrs %(count_users)d users',
        len(queryset)
    ) % {
        'count_users': len(queryset),
    }
    self.message_user(request, message)
make_users_as_non_active.short_description = _('Make selected users as non-active')


def make_users_as_active(self, request, queryset):
    """Admin action by selected users will be superusers."""
    for obj in queryset:
        obj.is_active = True
        obj.save()
    message = ungettext(
        'Succefull update parametrs one user',
        'Succefull update parametrs %(count_users)d users',
        len(queryset)
    ) % {
        'count_users': len(queryset),
    }
    self.message_user(request, message)
make_users_as_active.short_description = _('Make selected users as active')


def export_as_json(self, request, queryset):
    """Export seleted rows in admin in JSON-format."""

    response = HttpResponse(content_type='application/json')
    serializers.serialize('json', queryset, stream=response)
    return response
export_as_json.short_description = _('Export in JSON format')


def export_as_xml(self, request, queryset):
    """Export seleted rows in admin in XML-format."""

    response = HttpResponse(content_type='application/xml')
    serializers.serialize('xml', queryset, stream=response)
    return response
export_as_xml.short_description = _('Export in XML format')


def export_as_yaml(self, request, queryset):
    """Export seleted rows in admin in YAML-format."""

    response = HttpResponse(content_type='application/yaml')
    serializers.serialize('yaml', queryset, stream=response)
    return response
export_as_yaml.short_description = _('Export in YAML format')


def export_as_xls(self, request, queryset):
    """Export seleted rows in admin in XLS-format."""

export_as_xls.short_description = _('Export in Excel format (old version)')


def export_as_xlsx(self, request, queryset):
    """Export seleted rows in admin in XLSX-format."""

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "MyModel"

    row_num = 0

    columns = [
        # (u"ID", 15),
        (u"Title", 70),
        (u"Description", 70),
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            # obj.pk,
            obj.title,
            obj.description,
        ]
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.style.alignment.wrap_text = True

    wb.save(response)
    return response
export_as_xlsx.short_description = _('Export in Excel format (new version)')


def export_as_csv(self, request, queryset):
    """Export seleted rows in admin in CSV-format."""

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="somefilename.csv"'
    writer = csv.writer(response, csv.excel)
    response.write('\ufeff'.encode('utf8'))
    writer.writerow([
        smart_str("ID"),
        smart_str("Title"),
        smart_str("Description"),
    ])
    for obj in queryset:
        writer.writerow([
            smart_str(obj.pk),
            smart_str(obj.title),
            smart_str(obj.description),
        ])
    return response
export_as_csv.short_description = _('Export in CSV format')


def made_report(self, request, queryset):
    """Create report in PDF."""

    raise NotImplementedError
